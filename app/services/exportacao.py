"""Exportação da escala: planilha (.xlsx) dos grupos e calendário (.ics) do aluno.

Reaproveita os builders de contexto da UI (nenhuma lógica de escala nova aqui):
- Grupos → `estagios_dados.dados_grupos` (áreas → locais → ondas → membros + sem vaga).
- Calendário do aluno → sessões (prevista/cumprida) de cada alocação ativa.

Sem gate de perfil por enquanto (FASE 6): qualquer sessão pode baixar. O `.ics` usa
só a biblioteca padrão — cada sessão vira um VEVENT importável no Google Agenda etc.
"""
from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime, time, timezone
from io import BytesIO

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.aluno import Aluno, Matricula
from app.models.catalogo import Area
from app.models.ciclo import Ciclo
from app.models.enums import StatusAlocacao, StatusSessao
from app.models.escala import Alocacao, Sessao
from app.routers.ui import estagios_dados as ed
from app.services import aluno as aluno_service


# ------------------------------- utilidades -------------------------------
def slug(texto: str) -> str:
    """Nome de arquivo seguro (sem acento/espaço): 'João Silva' → 'joao_silva'."""
    base = unicodedata.normalize("NFKD", texto or "").encode("ascii", "ignore").decode()
    base = re.sub(r"[^A-Za-z0-9]+", "_", base).strip("_").lower()
    return base or "export"


# ------------------------------- Excel dos grupos -------------------------------
_CABECALHO = [
    "Área", "Local", "Unidade", "Dia", "Turno", "Horário", "Grupo",
    "Início", "Fim", "Capacidade", "Ocupação", "Aluno", "Ordenamento",
    "Origem", "Aviso",
]


def _hhmm(t: time | None) -> str:
    return t.strftime("%H:%M") if t else ""


def _ddmmaaaa(d: date | None) -> str:
    return d.strftime("%d/%m/%Y") if d else ""


def grupos_xlsx(db: Session, ciclo: Ciclo) -> bytes:
    """Planilha da escala: aba 'Grupos' (uma linha por aluno alocado) + aba 'Sem vaga'."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    dados = ed.dados_grupos(db, ciclo, "todas")
    wb = Workbook()

    cab_fill = PatternFill("solid", fgColor="1E293B")
    cab_font = Font(bold=True, color="FFFFFF")

    def _cabecalho(ws, colunas: list[str]) -> None:
        ws.append(colunas)
        for c in range(1, len(colunas) + 1):
            cel = ws.cell(row=1, column=c)
            cel.fill = cab_fill
            cel.font = cab_font
            cel.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

    def _largura(ws, colunas: list[str]) -> None:
        for i, _ in enumerate(colunas, start=1):
            comprimentos = [len(str(ws.cell(row=r, column=i).value or ""))
                            for r in range(1, ws.max_row + 1)]
            ws.column_dimensions[get_column_letter(i)].width = min(max(comprimentos + [10]) + 2, 45)

    # Aba principal: uma linha por membro.
    ws = wb.active
    ws.title = "Grupos"
    _cabecalho(ws, _CABECALHO)
    for area in dados["areas"]:
        for l in area["locais"]:
            horario = f"{_hhmm(l['hora_inicio'])}–{_hhmm(l['hora_fim'])}" if l["hora_inicio"] else ""
            for g in l["ondas"]:
                membros = g["membros"] or [None]  # grupo vazio ainda vira 1 linha (sem aluno)
                for m in membros:
                    ws.append([
                        area["nome"], l["campo"], l["unidade"] or "", l["dia"], l["turno"], horario,
                        f"Grupo {g['onda']}", _ddmmaaaa(g["data_inicio"]), _ddmmaaaa(g["data_fim"]),
                        g["cap"], g["ocup"],
                        m["nome"] if m else "(vazio)",
                        (m["ordenamento"] if m and m["ordenamento"] else ""),
                        ("montado" if m and m["fixado"] else ("auto" if m else "")),
                        (m["aviso"] if m and m["aviso"] else ""),
                    ])
    _largura(ws, _CABECALHO)

    # Aba secundária: alunos matriculados sem vaga no ciclo.
    ws2 = wb.create_sheet("Sem vaga")
    cols2 = ["Área", "Aluno", "Ordenamento", "Motivo"]
    _cabecalho(ws2, cols2)
    for area in dados["areas"]:
        for s in area["sem_vaga"]:
            ws2.append([area["nome"], s["nome"], s["ordenamento"] or "", s["motivo"]])
    _largura(ws2, cols2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------- Sessões do aluno (fonte comum p/ .ics e PDF) -------------------------------
_DIAS_SEMANA = ["segunda", "terça", "quarta", "quinta", "sexta", "sábado", "domingo"]


def _sessoes_aluno(db: Session, aluno_id: int) -> tuple[Aluno, list[dict]]:
    """Todas as sessões (prevista/cumprida) do aluno no ciclo, ordenadas por data/hora.
    Fonte única para o .ics e o PDF do calendário."""
    aluno = aluno_service.obter_model(db, aluno_id)  # 404 se não existir
    alocs = db.scalars(select(Alocacao).where(
        Alocacao.aluno_id == aluno_id,
        Alocacao.status != StatusAlocacao.cancelada)).all()
    itens: list[dict] = []
    for a in alocs:
        local = a.local
        area = db.get(Area, local.area_id) if local else None
        campo = local.campo if local else ""
        unidade = local.unidade if local and local.unidade else ""
        for s in db.scalars(select(Sessao).where(Sessao.alocacao_id == a.id)
                            .order_by(Sessao.data)).all():
            if s.status in (StatusSessao.cancelada, StatusSessao.remanejada):
                continue
            itens.append({
                "sessao_id": s.id,
                "data": s.data,
                "dia_semana": _DIAS_SEMANA[s.data.weekday()],
                "area_nome": area.nome if area else "Estágio",
                "area_cor": (area.cor if area else None) or "#64748b",
                "campo": campo, "unidade": unidade,
                "local_txt": " · ".join(x for x in (campo, unidade) if x),
                "hora_inicio": s.hora_inicio or (local.hora_inicio if local else None),
                "hora_fim": s.hora_fim or (local.hora_fim if local else None),
                "cumprida": s.status == StatusSessao.cumprida,
            })
    itens.sort(key=lambda x: (x["data"], x["hora_inicio"] or time.min))
    return aluno, itens


# ------------------------------- Calendário do aluno (.ics) -------------------------------
def _escape_ics(texto: str) -> str:
    """Escapa caracteres especiais do iCalendar (RFC 5545 §3.3.11)."""
    return (texto.replace("\\", "\\\\").replace(";", "\\;")
            .replace(",", "\\,").replace("\n", "\\n"))


def _fold(linha: str) -> str:
    """Dobra linhas > 75 octetos com CRLF + espaço (RFC 5545 §3.1)."""
    if len(linha.encode("utf-8")) <= 75:
        return linha
    partes, atual = [], ""
    for ch in linha:
        if len((atual + ch).encode("utf-8")) > 75:
            partes.append(atual)
            atual = " " + ch  # continuação começa com espaço
        else:
            atual += ch
    partes.append(atual)
    return "\r\n".join(partes)


def calendario_aluno_ics(db: Session, aluno_id: int) -> tuple[str, str]:
    """Devolve (nome_arquivo, conteúdo_ics) com todas as sessões (prevista/cumprida)
    de todas as áreas do aluno no ciclo. Cada sessão = um VEVENT."""
    aluno, itens = _sessoes_aluno(db, aluno_id)  # 404 se não existir
    agora = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")

    linhas = [
        "BEGIN:VCALENDAR", "VERSION:2.0",
        "PRODID:-//UFCSPA//Gestao Estagios Fono//PT-BR",
        "CALSCALE:GREGORIAN", "METHOD:PUBLISH",
        f"X-WR-CALNAME:{_escape_ics('Estágios — ' + aluno.nome.strip())}",
    ]

    for it in itens:
        ini, fim = it["hora_inicio"], it["hora_fim"]
        sufixo = " (realizado)" if it["cumprida"] else ""
        linhas.append("BEGIN:VEVENT")
        linhas.append(f"UID:sessao-{it['sessao_id']}@estagios.ufcspa.edu.br")
        linhas.append(f"DTSTAMP:{agora}")
        if ini and fim:
            linhas.append(f"DTSTART:{it['data'].strftime('%Y%m%d')}T{ini.strftime('%H%M%S')}")
            linhas.append(f"DTEND:{it['data'].strftime('%Y%m%d')}T{fim.strftime('%H%M%S')}")
        else:  # sem horário → evento de dia inteiro
            linhas.append(f"DTSTART;VALUE=DATE:{it['data'].strftime('%Y%m%d')}")
        linhas.append(_fold(f"SUMMARY:{_escape_ics(it['area_nome'] + sufixo)}"))
        if it["local_txt"]:
            linhas.append(_fold(f"LOCATION:{_escape_ics(it['local_txt'])}"))
        linhas.append(_fold(f"DESCRIPTION:{_escape_ics('Estágio de ' + it['area_nome'] + ' — ' + aluno.nome.strip())}"))
        linhas.append("END:VEVENT")

    linhas.append("END:VCALENDAR")
    conteudo = "\r\n".join(linhas) + "\r\n"
    nome = f"calendario_{slug(aluno.nome)}.ics"
    return nome, conteudo


# ------------------------------- PDF (reportlab) -------------------------------
def _cor(hex_str: str | None):
    """'#rrggbb' → reportlab Color; cinza como fallback."""
    from reportlab.lib.colors import HexColor
    try:
        return HexColor(hex_str) if hex_str else HexColor("#64748b")
    except Exception:
        return HexColor("#64748b")


def _estilos():
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    ss = getSampleStyleSheet()
    ss.add(ParagraphStyle("TituloDoc", parent=ss["Title"], fontSize=16, spaceAfter=2, alignment=TA_LEFT))
    ss.add(ParagraphStyle("Sub", parent=ss["Normal"], fontSize=9, textColor=_cor("#64748b"), spaceAfter=10))
    ss.add(ParagraphStyle("Area", parent=ss["Heading2"], fontSize=12, spaceBefore=12, spaceAfter=4))
    ss.add(ParagraphStyle("Local", parent=ss["Heading3"], fontSize=10, spaceBefore=6, spaceAfter=2))
    ss.add(ParagraphStyle("Cel", parent=ss["Normal"], fontSize=8, leading=10))
    return ss


def _estilo_tabela(cor_cabecalho):
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), cor_cabecalho),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
    ])


def grupos_pdf(db: Session, ciclo: Ciclo) -> bytes:
    """PDF da escala: por área → local → grupos (ondas) com datas e membros."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table

    dados = ed.dados_grupos(db, ciclo, "todas")
    ss = _estilos()
    story = [
        Paragraph("Escala de Estágios — Grupos", ss["TituloDoc"]),
        Paragraph(f"Ciclo {ciclo.data_inicio.year} · gerado em {date.today().strftime('%d/%m/%Y')}", ss["Sub"]),
    ]

    if not dados["areas"]:
        story.append(Paragraph("Nenhum grupo gerado ainda.", ss["Normal"]))
    for area in dados["areas"]:
        cor = _cor(area["cor"])
        story.append(Paragraph(f'<font color="{area["cor"] or "#64748b"}">■</font> {area["nome"]}', ss["Area"]))
        for l in area["locais"]:
            horario = f"{_hhmm(l['hora_inicio'])}–{_hhmm(l['hora_fim'])}" if l["hora_inicio"] else "—"
            cab = f"<b>{l['campo']}</b>"
            if l["unidade"]:
                cab += f' · {l["unidade"]}'
            cab += f' · {l["dia"]} · {l["turno"]} · {horario} · cap. {l["cap"]}'
            story.append(Paragraph(cab, ss["Local"]))
            linhas = [["Grupo", "Início", "Fim", "Ocup.", "Aluno", "Ord.", "Origem", "Aviso"]]
            for g in l["ondas"]:
                membros = g["membros"] or [None]
                for i, m in enumerate(membros):
                    linhas.append([
                        f"G{g['onda']}" if i == 0 else "",
                        _ddmmaaaa(g["data_inicio"]) if i == 0 else "",
                        _ddmmaaaa(g["data_fim"]) if i == 0 else "",
                        f"{g['ocup']}/{g['cap']}" if i == 0 else "",
                        Paragraph(m["nome"], ss["Cel"]) if m else "(vazio)",
                        str(m["ordenamento"]) if m and m["ordenamento"] else "",
                        ("montado" if m and m["fixado"] else ("auto" if m else "")),
                        Paragraph(m["aviso"], ss["Cel"]) if m and m["aviso"] else "",
                    ])
            larguras = [1.3*cm, 2.1*cm, 2.1*cm, 1.4*cm, 7.5*cm, 1.2*cm, 2*cm, 6*cm]
            t = Table(linhas, colWidths=larguras, repeatRows=1)
            t.setStyle(_estilo_tabela(cor))
            story.append(t)
            story.append(Spacer(1, 6))
        for s in area["sem_vaga"]:
            story.append(Paragraph(
                f'<font color="#c0392b">Sem vaga:</font> {s["nome"]}'
                + (f' ({s["ordenamento"]}º)' if s["ordenamento"] else "") + f' — {s["motivo"]}',
                ss["Cel"]))

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.2*cm, rightMargin=1.2*cm,
                            topMargin=1.2*cm, bottomMargin=1.2*cm,
                            title=f"Grupos {ciclo.data_inicio.year}")
    doc.build(story)
    return buf.getvalue()


def calendario_aluno_pdf(db: Session, aluno_id: int) -> tuple[str, bytes]:
    """PDF do calendário do aluno: progresso por área + lista de encontros por data."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table

    from app.routers.ui.aluno_dados import montar_encontros
    aluno, itens = _sessoes_aluno(db, aluno_id)
    resumo = montar_encontros(db, aluno_id, None)["resumo"]
    ss = _estilos()

    story = [
        Paragraph(f"Calendário de Estágios — {aluno.nome.strip()}", ss["TituloDoc"]),
        Paragraph(f"Matrícula {aluno.matricula} · {aluno.semestre}º semestre · "
                  f"gerado em {date.today().strftime('%d/%m/%Y')}", ss["Sub"]),
    ]

    # Progresso por área.
    if resumo:
        story.append(Paragraph("Progresso por área", ss["Area"]))
        linhas = [["Área", "Encontros", "%", "Situação"]]
        for r in resumo:
            linhas.append([Paragraph(r["area_nome"], ss["Cel"]), f"{r['feitos']}/{r['total']}",
                           f"{r['pct']}%", "concluída" if r["concluida"] else "em andamento"])
        t = Table(linhas, colWidths=[9*cm, 3*cm, 2*cm, 4*cm], repeatRows=1)
        t.setStyle(_estilo_tabela(_cor("#0ea5e9")))
        story.append(t)
        story.append(Spacer(1, 10))

    # Encontros por data.
    story.append(Paragraph("Encontros", ss["Area"]))
    if itens:
        linhas = [["Data", "Dia", "Horário", "Área", "Local", "Situação"]]
        for it in itens:
            horario = (f"{_hhmm(it['hora_inicio'])}–{_hhmm(it['hora_fim'])}"
                       if it["hora_inicio"] else "—")
            linhas.append([
                it["data"].strftime("%d/%m/%Y"), it["dia_semana"], horario,
                Paragraph(it["area_nome"], ss["Cel"]), Paragraph(it["local_txt"], ss["Cel"]),
                "realizado" if it["cumprida"] else "previsto",
            ])
        t = Table(linhas, colWidths=[2.4*cm, 2*cm, 2.6*cm, 4.5*cm, 5*cm, 2.3*cm], repeatRows=1)
        t.setStyle(_estilo_tabela(_cor("#0ea5e9")))
        story.append(t)
    else:
        story.append(Paragraph("Sem encontros na escala — gere a escala primeiro.", ss["Normal"]))

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            title=f"Calendario {aluno.nome.strip()}")
    doc.build(story)
    nome = f"calendario_{slug(aluno.nome)}.pdf"
    return nome, buf.getvalue()
